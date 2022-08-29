import datetime

import qrcode
from openpyxl import load_workbook, Workbook

file = 'C:/Users/12267/Desktop/数据源.xlsx'
book = load_workbook('C:/Users/12267/Desktop/销项发票信息03.xlsx')
# wb = load_workbook(file)
# sheet = wb.active  # 打开默认激活的工作表
x = 0

# print(sheet.max_row)
# for r in range()
ws = book.active
print(ws.max_row)
count = 0
while ws.max_row != 1:
    count += 1
    for row in range(ws.max_row)[1:]:
        ws.delete_rows(row + 1)
        print('删除第', row, '行')
    print(str(datetime.datetime.now()) + ' 删除第', count, '次')
print(ws.max_row)
book.save('C:/Users/12267/Desktop/销项发票信息03.xlsx')
