import datetime

import qrcode
from openpyxl import load_workbook, Workbook
import pandas as pd

file = 'C:/Users/12267/Desktop/数据源.xlsx'
wb = load_workbook(file)
sheet = wb.active  # 打开默认激活的工作表
x = 0
'''
df = pd.read_excel('C:/Users/12267/Desktop/数据源.xlsx')
print(df)
for i in range(len(df)):
    context = dict(df.iloc[i])
    # print(context)
    fileName = context['姓名']
    fileNum = context['编号']
    img = qrcode.make(context)
    filename = './file/' + str(fileNum) + '_' + fileName + '.png'
    img.save(filename)
    x += 1
'''



for r in range(2, sheet.max_row + 1):
    # n列数据生成条码
    var1 = sheet.cell(column=1, row=r).value
    var2 = sheet.cell(column=2, row=r).value
    var = str(var1) + '_' + var2
    img = qrcode.make(var)
    filename = str(var) + '.png'
    img.save('./file/' + filename)
    x += 1
wb.save(file)
print('执行成功！共计%s条！' % str(x))
