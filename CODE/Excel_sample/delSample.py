from openpyxl import Workbook, load_workbook

book = load_workbook('C:/Users/12267/Desktop/销项发票信息.xlsx')
ws = book.active
row_number = 0
deleteCom = ['E36','E52','E82','E99','E100','E101','E102','E103','E107','E108','E109','E111','E112','E113','E114','E115','E116','E117','E118','E119','E120','E121','E122','E123']
deleteCom1 = ['E124']

for id in deleteCom1:
    for row in ws.iter_rows():  # 迭代遍历每行
        if row[0].value == id:#给定的条件，读者可根据自身需求自定义
            row_number = row[1].row  # 关键步骤！获得当前行的行号！
            ws.delete_rows(row_number)
            print(id, '删除', row_number,'行')
    # print('row[1]: ', row[1])
    # print('row_number:', row_number)

book.save('C:/Users/12267/Desktop/销项发票信息.xlsx')
