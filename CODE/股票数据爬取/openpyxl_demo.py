import random
import time
import xlwt
import method
from openpyxl import Workbook, load_workbook

book = Workbook()
print('工作簿创建完成')
url = 'https://xueqiu.com/hq'
start = time.time()
print('开始爬取数据')
###找到有哪些页面需要爬取
hq_url = 'https://xueqiu.com/hq'
codes = method.get_code(hq_url=hq_url)  # [类型代号，股票类型]
print('已拿到', len(codes), '个行业代号')
last = ['1660795197446', '1660795071180', '1660795086940', '1660795115696', '1660795131328', '1660795196330',
        '1660795196960', '1660795247290', '1660795679516', '1660877397824', '1660876865775', '1660876840948']
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'
}
# book = xlwt.Workbook(encoding='utf-8', style_compression=0)  # 创建一个文件，设置编码格式，且不压缩
for code in codes:
    # 创建表格
    sheet = book.create_sheet(code[0])  # 在文件中添加一个表
    col = ('股票代码', '股票名称', '当前价', '涨跌额', '涨跌幅', '年初至今', '成交量', '成交额', '换手率', '市盈率(TTM)', '股息率', '市值')  # 写明列数据
    for i in range(0, 12):
        sheet.write(0, i, col[i])  # 0行 i列， 写入col[i]
    stockList = method.re_getStockList(code=code[1], headers=headers, last_url=random.choice(last),
                                       count=method.get_stockCount(code=code[1], headers=headers, last_url=random.choice(last)))
    print(code[0], '行业数据爬取成功！')
    for i in range(0, len(stockList)):  # n组数据
        data = stockList[i]
        for j in range(0, len(col)):  # 每组数据有8列
            sheet.write(i + 1, j, data[j])
savePath = 'C:/Users/12267/Desktop/股票.xlsx'
book.save(savePath)
end = time.time()
print('数据已导入Excel,共', len(codes), '条数据!')
print('用时', round(end - start, 4), '秒')