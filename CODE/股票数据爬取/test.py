import random
import time
import xlwt
import method
from openpyxl import Workbook, load_workbook
import pprint

# book = Workbook('')
# print('工作簿创建完成')
# cols = ('股票代码', '股票名称', '当前价', '涨跌额', '涨跌幅', '年初至今', '成交量', '成交额', '换手率', '市盈率(TTM)', '股息率', '市值')
# for i in range(1, 11):
#     sheet = book.create_sheet('sheet' + str(i))
#     for j in range(1, 13):
#         sheet.cell(1, j).value = cols[j - 1]
#     if i % 5 == 0:
#         sheet.sheet_properties.tabColor = '1072BA'
# book.save('C:/Users/12267/Desktop/股票.xlsx')
start = time.time()
book = load_workbook('C:/Users/12267/Desktop/销项发票信息02.xlsx')
ws = book.active
output = 0
codeList = ['E36', 'E52', 'E82', 'E99', 'E100', 'E101', 'E102', 'E103', 'E107', 'E108', 'E109', 'E111', 'E112', 'E113',
            'E114', 'E115', 'E116', 'E117', 'E118', 'E119', 'E120', 'E121', 'E122', 'E123']
for i in range(2, len(tuple(ws.rows))):
    for code in codeList:
        if ws.cell(row=i, column=1).value == code:
            ws.delete_rows(i)
            print('删除', code, '的行数：', i)
# print('平均销项发票税额:', output)
print('需要删除的企业代号:', codeList)
book.save('C:/Users/12267/Desktop/销项发票信息02.xlsx')
end = time.time()
print('用时', (end - start) / 1000, '分钟')
